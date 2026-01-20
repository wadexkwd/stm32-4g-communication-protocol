#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
MQTT消息监听器
功能：
- 连接到指定的MQTT服务器
- 订阅主题：up/861197065268692
- 打印接收到的消息
- 支持自动重连
- 格式化输出传感器数据
- 将数据写入Excel文件
"""

import paho.mqtt.client as mqtt
import json
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# =============================================================================
# 配置参数
# =============================================================================
MQTT_BROKER = "120.27.250.30"  # MQTT服务器地址
MQTT_PORT = 1883  # MQTT端口
MQTT_USERNAME = ""  # MQTT用户名（不需要认证）
MQTT_PASSWORD = ""  # MQTT密码（不需要认证）
IMEI = "861197065268692"  # 设备IMEI号
MQTT_TOPIC = f"up/{IMEI}"  # 订阅的主题
CLIENT_ID = f"windows_listener_{IMEI}"  # 客户端ID，确保唯一性
EXCEL_FILE = "sensor_data.xlsx"  # 输出的Excel文件名

# 传感器数据字段定义（按照协议顺序）
FIELD_ORDER = [
    'timestamp',
    'version',
    'packet_order',
    'accel_x',
    'accel_y',
    'accel_z',
    'gyro_x',
    'gyro_y',
    'gyro_z',
    'angle_x',
    'angle_y',
    'angle_z',
    'attitude1',
    'attitude2',
    'pressure',
    'altitude',
    'longitude',
    'latitude'
]

# 字段名映射（用于Excel表头）
FIELD_NAMES = {
    'timestamp': '时间戳',
    'version': '版本',
    'packet_order': '包序',
    'accel_x': '加速度X',
    'accel_y': '加速度Y',
    'accel_z': '加速度Z',
    'gyro_x': '角速度X',
    'gyro_y': '角速度Y',
    'gyro_z': '角速度Z',
    'angle_x': '角度X',
    'angle_y': '角度Y',
    'angle_z': '角度Z',
    'attitude1': '姿态角1',
    'attitude2': '姿态角2',
    'pressure': '气压',
    'altitude': '高度',
    'longitude': '经度',
    'latitude': '纬度'
}

# =============================================================================
# Excel文件操作函数
# =============================================================================
def init_excel():
    """初始化Excel文件"""
    if not os.path.exists(EXCEL_FILE):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "传感器数据"
            
            # 写入表头
            for col, field in enumerate(FIELD_ORDER, 1):
                ws[get_column_letter(col) + '1'] = FIELD_NAMES.get(field, field)
                
            wb.save(EXCEL_FILE)
            print(f"✅ Excel文件已创建: {EXCEL_FILE}")
        except Exception as e:
            print(f"❌ 创建Excel文件失败: {e}")

def write_to_excel(data_list):
    """将数据写入Excel文件"""
    max_retries = 3
    retry_delay = 0.5
    
    for attempt in range(max_retries):
        try:
            # 检查文件是否存在且有效
            if os.path.exists(EXCEL_FILE):
                try:
                    # 尝试加载文件
                    wb = load_workbook(EXCEL_FILE)
                    ws = wb.active
                except Exception as e:
                    print(f"❌ Excel文件损坏，将创建新文件: {e}")
                    # 删除损坏的文件
                    try:
                        os.remove(EXCEL_FILE)
                    except Exception as remove_e:
                        print(f"❌ 无法删除损坏的文件: {remove_e}")
                        time.sleep(retry_delay)
                        continue
                    # 创建新文件
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "传感器数据"
                    
                    # 写入表头
                    for col, field in enumerate(FIELD_ORDER, 1):
                        ws[get_column_letter(col) + '1'] = FIELD_NAMES.get(field, field)
            else:
                # 文件不存在，创建新文件
                wb = Workbook()
                ws = wb.active
                ws.title = "传感器数据"
                
                # 写入表头
                for col, field in enumerate(FIELD_ORDER, 1):
                    ws[get_column_letter(col) + '1'] = FIELD_NAMES.get(field, field)
        
            # 找到下一个可用行
            next_row = ws.max_row + 1
            
            # 写入数据 
            for data in data_list:
                for col, field in enumerate(FIELD_ORDER, 1):
                    value = data.get(field, '')
                    ws[get_column_letter(col) + str(next_row)] = value
                next_row += 1
                
            wb.save(EXCEL_FILE)
            print(f"✅ 已写入 {len(data_list)} 条数据到Excel文件")
            return  # 成功写入，返回
        except Exception as e:
            print(f"❌ 写入Excel文件失败（尝试 {attempt + 1}/{max_retries}）: {e}")
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
            else:
                print(f"❌ 多次尝试失败，将数据保存到临时文件")
                # 尝试保存到临时文件
                temp_file = f"{EXCEL_FILE}.tmp"
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "传感器数据"
                    
                    # 写入表头
                    for col, field in enumerate(FIELD_ORDER, 1):
                        ws[get_column_letter(col) + '1'] = FIELD_NAMES.get(field, field)
                        
                    # 写入数据
                    next_row = 2
                    for data in data_list:
                        for col, field in enumerate(FIELD_ORDER, 1):
                            value = data.get(field, '')
                            ws[get_column_letter(col) + str(next_row)] = value
                        next_row += 1
                        
                    wb.save(temp_file)
                    print(f"✅ 数据已保存到临时文件: {temp_file}")
                except Exception as temp_e:
                    print(f"❌ 保存临时文件失败: {temp_e}")

# =============================================================================
# 数据格式化输出函数
# =============================================================================
def format_sensor_data(data_list):
    """格式化传感器数据输出"""
    if not data_list:
        return
        
    # 计算字符显示宽度（中文字符占2个宽度，西文字符占1个）
    def get_display_width(s):
        width = 0
        for char in str(s):
            if '\u4e00' <= char <= '\u9fff' or char in '。，；：！？“”‘’（）《》【】':
                width += 2
            else:
                width += 1
        return width
    
    # 填充字符串到指定显示宽度
    def pad_str(s, target_width):
        current_width = get_display_width(s)
        padding = target_width - current_width
        return str(s) + ' ' * padding
    
    # 定义每个字段的显示宽度（考虑中文字符）
    field_widths = [
        22,  # 时间戳：2026-01-16 08:16:111001 (19个字符，西文)
        6,   # 版本 (中文2字，4 + 2)
        6,   # 包序 (中文2字)
        10,  # 加速度X (中文4字，8 + 2)
        10,  # 加速度Y
        10,  # 加速度Z
        10,  # 角速度X
        10,  # 角速度Y
        10,  # 角速度Z
        8,   # 角度X (中文3字，6 + 2)
        8,   # 角度Y
        8,   # 角度Z
        10,  # 姿态角1 (中文4字)
        10,  # 姿态角2
        6,   # 气压 (中文2字)
        14,  # 高度 (数值可能较长)
        8,   # 经度 (中文2字)
        8    # 纬度 (中文2字)
    ]
    
    # 计算总宽度（加上字段之间的分隔空格）
    total_width = sum(field_widths) + len(field_widths) + 2
    
    print("\n" + "=" * total_width)
    print("传感器数据:")
    
    # 打印表头
    header = "  "
    headers = [
        '时间戳', '版本', '包序', '加速度X', '加速度Y', '加速度Z',
        '角速度X', '角速度Y', '角速度Z', '角度X', '角度Y', '角度Z',
        '姿态角1', '姿态角2', '气压', '高度', '经度', '纬度'
    ]
    for i, h in enumerate(headers):
        header += pad_str(h, field_widths[i]) + " "
    print(header)
    
    print("-" * total_width)
    
    # 打印数据行
    for data in data_list:
        row = "  "
        values = [
            str(data.get('timestamp', '')), str(data.get('version', '')),
            str(data.get('packet_order', '')), str(data.get('accel_x', '')),
            str(data.get('accel_y', '')), str(data.get('accel_z', '')),
            str(data.get('gyro_x', '')), str(data.get('gyro_y', '')),
            str(data.get('gyro_z', '')), str(data.get('angle_x', '')),
            str(data.get('angle_y', '')), str(data.get('angle_z', '')),
            str(data.get('attitude1', '')), str(data.get('attitude2', '')),
            str(data.get('pressure', '')), str(data.get('altitude', '')),
            str(data.get('longitude', '')), str(data.get('latitude', ''))
        ]
        for i, v in enumerate(values):
            row += pad_str(v, field_widths[i]) + " "
        print(row)
    
    print("=" * total_width)
    print(f"数据条数: {len(data_list)}\n")

# =============================================================================
# MQTT事件处理函数
# =============================================================================
def on_connect(client, userdata, flags, rc, properties):
    """连接成功回调函数 - API Version 2"""
    if rc == 0:
        print("✅ MQTT连接成功")
        # 订阅主题
        client.subscribe(MQTT_TOPIC)
        print(f"✅ 已订阅主题: {MQTT_TOPIC}")
    else:
        print(f"❌ MQTT连接失败，错误码: {rc}")

def on_message(client, userdata, msg):
    """消息接收回调函数"""
    try:
        # 解码消息
        payload = msg.payload.decode('utf-8')
        
        # 尝试解析JSON格式消息
        try:
            data = json.loads(payload)
            
            # 如果是传感器数据列表
            if isinstance(data, list):
                print(f"\n📩 收到 {len(data)} 条传感器数据")
                format_sensor_data(data)
                write_to_excel(data)
            else:
                # 其他类型的消息（如心跳包、配置参数等）
                print(f"\n📩 收到消息:")
                print(f"   主题: {msg.topic}")
                print(f"   内容: {payload}")
                
        except json.JSONDecodeError:
            print(f"\n📩 收到非JSON格式消息:")
            print(f"   主题: {msg.topic}")
            print(f"   内容: {payload}")
            
    except Exception as e:
        print(f"\n❌ 消息处理失败: {e}")

def on_disconnect(client, userdata, rc, properties, reason_code):
    """断开连接回调函数 - API Version 2"""
    # 格式化显示断开连接的原因
    if hasattr(rc, 'is_disconnect_packet_from_server'):
        if rc.is_disconnect_packet_from_server:
            print("🔌 MQTT连接已断开（服务器主动断开）")
        else:
            print("🔌 MQTT连接已断开（客户端主动断开）")
    else:
        print(f"🔌 MQTT连接已断开，错误码: {rc}")
    
    print("⏳ 正在尝试重连...")

def on_log(client, userdata, level, buf):
    """日志回调函数（可选）"""
    # 可以在这里添加详细的日志记录
    # print(f"📝 日志: {buf}")
    pass

# =============================================================================
# 主函数
# =============================================================================
def main():
    """主函数"""
    # 初始化Excel文件
    init_excel()
    
    print("=" * 60)
    print("MQTT消息监听器")
    print("=" * 60)
    print(f"服务器地址: {MQTT_BROKER}:{MQTT_PORT}")
    print(f"订阅主题: {MQTT_TOPIC}")
    print(f"客户端ID: {CLIENT_ID}")
    print(f"Excel文件: {EXCEL_FILE}")
    print("=" * 60)
    
    # 创建MQTT客户端 - 使用最新的API版本
    client = mqtt.Client(client_id=CLIENT_ID, callback_api_version=mqtt.CallbackAPIVersion.VERSION2)
    
    # 设置回调函数
    client.on_connect = on_connect
    client.on_message = on_message
    client.on_disconnect = on_disconnect
    client.on_log = on_log
    
    # 设置认证信息
    if MQTT_USERNAME:
        client.username_pw_set(MQTT_USERNAME, MQTT_PASSWORD)
    
    # 设置连接参数
    client.reconnect_delay_set(min_delay=1, max_delay=60)
    client.keepalive = 120  # 心跳间隔
    
    # 连接到MQTT服务器
    try:
        print("📡 正在连接到MQTT服务器...")
        client.connect(MQTT_BROKER, MQTT_PORT)
    except Exception as e:
        print(f"❌ 连接失败: {e}")
        return
    
    # 保持连接并持续监听
    try:
        print("\n🚀 开始监听消息（按 Ctrl+C 停止）")
        print("-" * 60)
        client.loop_forever()
    except KeyboardInterrupt:
        print("\n\n🛑 用户中断")
    except Exception as e:
        print(f"\n❌ 程序异常: {e}")
    finally:
        client.disconnect()
        print("🔌 已断开MQTT连接")

if __name__ == "__main__":
    main()
