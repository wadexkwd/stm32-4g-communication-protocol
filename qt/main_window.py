#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
应急跌落事件监控系统 - Qt主窗口
功能：
- 集成各个功能模块，提供完整的用户界面
- 处理用户交互和事件响应
- 协调各个模块之间的通信
"""

import sys
import os
import time
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QGroupBox, QLabel, QLineEdit, QPushButton, QComboBox,
    QTextEdit, QTableWidget, QTableWidgetItem, QDateTimeEdit, QFileDialog,
    QMessageBox, QSplitter, QHeaderView, QProgressBar
)
from PySide6.QtCore import Qt, QTimer, QDateTime, QUrl
from PySide6.QtGui import QFont, QColor, QPainter, QIcon, QPixmap
from PySide6.QtCharts import QChart, QChartView, QLineSeries, QValueAxis
from PySide6.QtWebEngineWidgets import QWebEngineView
from PySide6.QtWebEngineCore import QWebEnginePage
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config import (
    FIELD_ORDER, FIELD_NAMES, EVENT_TYPES, FIELD_UNITS, FIELD_CATEGORIES,
    DATABASE_FILE
)
from database_manager import DatabaseManager
from mqtt_thread import MqttThread
from log_window import LogWindow
from utils import StreamRedirector

# =============================================================================
# 配置参数
# =============================================================================
# 软件版本
APP_VERSION = "v1.2"

# 地图配置
MAP_API_DAILY_LIMIT = 1000  # 地图API每日调用次数限制
MAP_UPDATE_DISTANCE_THRESHOLD = 50  # 地图更新距离阈值（米）
MAP_ZOOM_LEVEL = 15  # 地图显示缩放级别

# 日志配置
MAX_LOG_FILE_SIZE = 5 * 1024 * 1024  # 日志文件最大大小（5MB）

# 图表配置
CHART_DISPLAY_DURATION = 10  # 图表显示时间范围（秒）
CHART_MAX_DATA_POINTS = 3000  # 图表最大数据点数量
CHART_Y_AXIS_MARGIN = 0.1  # 图表Y轴边距比例（10%）

# 数据加载配置
DATA_LOAD_TIME_RANGE = 300  # 数据加载时间范围（秒）

# MQTT配置
MQTT_RECONNECT_DELAY = 1000  # MQTT重连延迟（毫秒）

# 其他配置
WINDOW_GEOMETRY = (100, 100, 1400, 900)  # 窗口几何尺寸（x, y, width, height）

# =============================================================================
# 主窗口类
# =============================================================================
class MainWindow(QMainWindow):
    """主窗口类"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"应急跌落事件监控系统{APP_VERSION}")
        self.setGeometry(*WINDOW_GEOMETRY)
        
        # 初始化时间记录变量
        self.start_time = time.time()
        
        # 初始化数据库
        self.db_manager = DatabaseManager()
        
        # 初始化MQTT线程
        self.mqtt_thread = None
        
        # 已订阅的设备列表
        self.subscribed_devices = []
        
        # 原始数据log文件相关属性
        self.log_file = None
        self.log_file_path = None
        self.log_file_size = 0
        self.log_file_counter = 0
        self.max_log_file_size = MAX_LOG_FILE_SIZE
        
        # 创建UI
        self.create_ui()
        
        # 加载已订阅的设备
        self.load_subscribed_devices()
        
        # 初始化日志窗口
        self.log_window = LogWindow()
        
        # 重定向标准输出和错误
        self.redirect_stdout_stderr()
        
        # 打印初始化信息
        print("应急跌落事件监控系统启动成功")
        print(f"当前时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    def create_ui(self):
        """创建用户界面"""
        # 中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局
        main_layout = QVBoxLayout(central_widget)
        
        # 顶部控制区 - 垂直方向缩小，不超过4行
        control_group = QGroupBox("控制区域")
        control_layout = QHBoxLayout(control_group)
        
        # IMEI输入框
        self.imei_label = QLabel("设备IMEI:")
        self.imei_edit = QLineEdit()
        self.imei_edit.setPlaceholderText("请输入设备IMEI号")
        self.imei_edit.setFixedWidth(200)
        
        # 订阅按钮
        self.subscribe_btn = QPushButton("订阅设备")
        self.subscribe_btn.clicked.connect(self.subscribe_device)
        
        # 设备选择下拉框
        self.device_combo = QComboBox()
        self.device_combo.setFixedWidth(200)
        self.device_combo.currentTextChanged.connect(self.switch_device)
        
        # 事件类型筛选下拉框
        self.event_filter_combo = QComboBox()
        self.event_filter_combo.addItem("全部事件")
        for event_type in EVENT_TYPES.values():
            self.event_filter_combo.addItem(event_type)
        self.event_filter_combo.currentTextChanged.connect(self.filter_data_by_event)
        
        # 日志按钮
        self.log_btn = QPushButton("查看日志")
        self.log_btn.clicked.connect(self.show_log_window)
        
        # 状态标签
        self.status_label = QLabel("状态: 未连接")
        self.status_label.setStyleSheet("color: gray; font-weight: bold;")
        
        # 连接控制区组件
        control_layout.addWidget(self.imei_label)
        control_layout.addWidget(self.imei_edit)
        control_layout.addWidget(self.subscribe_btn)
        control_layout.addWidget(QLabel("已订阅设备:"))
        control_layout.addWidget(self.device_combo)
        control_layout.addWidget(QLabel("事件类型:"))
        control_layout.addWidget(self.event_filter_combo)
        control_layout.addWidget(self.log_btn)
        control_layout.addStretch()
        control_layout.addWidget(self.status_label)
        
        main_layout.addWidget(control_group)
        
        # 主内容区 - 垂直布局，大小可手动调整
        content_splitter = QSplitter(Qt.Vertical)
        
        # 解析数据展示区
        parsed_widget = QWidget()
        parsed_layout = QVBoxLayout(parsed_widget)
        
        # 标签页控件
        self.tab_widget = QTabWidget()
        
        # 数据总览标签页
        overview_tab = QWidget()
        overview_layout = QVBoxLayout(overview_tab)
        
        self.overview_table = QTableWidget()
        self.overview_table.setColumnCount(len(FIELD_ORDER))
        self.overview_table.setHorizontalHeaderLabels([FIELD_NAMES[f] for f in FIELD_ORDER])
        # 设置列宽调整模式为可交互（支持手动调节）
        self.overview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        # 设置默认列宽
        self.overview_table.horizontalHeader().setDefaultSectionSize(100)
        self.overview_table.setAlternatingRowColors(True)
        # 监听滚动条事件，实现滚动加载更多数据
        self.overview_table.verticalScrollBar().valueChanged.connect(self.on_table_scroll)
        # 用于记录已加载的最早时间
        self.loaded_earliest_time = QDateTime.currentDateTime()
        
        overview_layout.addWidget(self.overview_table)
        self.tab_widget.addTab(overview_tab, "数据总览")
        
        # 分类展示标签页
        for category, fields in FIELD_CATEGORIES.items():
            category_tab = QWidget()
            category_layout = QVBoxLayout(category_tab)
            
            if category == "加速度":
                # 加速度标签页：添加图表展示
                chart_layout = QHBoxLayout()
                
                # 创建加速度图表
                self.accel_chart = QChart()
                self.accel_chart.setTitle("加速度随时间变化曲线")
                self.accel_chart.setAnimationOptions(QChart.NoAnimation)  # 禁用动画，提高性能
                
                # 创建三个坐标轴（X轴为时间，Y轴为加速度值）
                self.accel_axis_x = QValueAxis()
                self.accel_axis_x.setTitleText("时间 (s)")
                self.accel_axis_x.setLabelFormat("%.1f")
                self.accel_axis_x.setRange(0, 10)  # 显示最近10秒数据
                
                self.accel_axis_y = QValueAxis()
                self.accel_axis_y.setTitleText("加速度 (mg)")
                self.accel_axis_y.setLabelFormat("%.0f")
                self.accel_axis_y.setRange(-2000, 2000)  # 合理的加速度范围
                
                # 创建三个系列（X、Y、Z轴）
                self.accel_series_x = QLineSeries()
                self.accel_series_x.setName("加速度X (mg)")
                self.accel_series_x.setColor(QColor("red"))
                
                self.accel_series_y = QLineSeries()
                self.accel_series_y.setName("加速度Y (mg)")
                self.accel_series_y.setColor(QColor("green"))
                
                self.accel_series_z = QLineSeries()
                self.accel_series_z.setName("加速度Z (mg)")
                self.accel_series_z.setColor(QColor("blue"))
                
                # 添加系列和坐标轴到图表
                self.accel_chart.addSeries(self.accel_series_x)
                self.accel_chart.addSeries(self.accel_series_y)
                self.accel_chart.addSeries(self.accel_series_z)
                
                self.accel_chart.addAxis(self.accel_axis_x, Qt.AlignBottom)
                self.accel_chart.addAxis(self.accel_axis_y, Qt.AlignLeft)
                
                self.accel_series_x.attachAxis(self.accel_axis_x)
                self.accel_series_x.attachAxis(self.accel_axis_y)
                
                self.accel_series_y.attachAxis(self.accel_axis_x)
                self.accel_series_y.attachAxis(self.accel_axis_y)
                
                self.accel_series_z.attachAxis(self.accel_axis_x)
                self.accel_series_z.attachAxis(self.accel_axis_y)
                
                # 创建图表视图
                chart_view = QChartView(self.accel_chart)
                chart_view.setRenderHint(QPainter.Antialiasing)
                chart_view.setMinimumHeight(300)
                
                chart_layout.addWidget(chart_view)
                category_layout.addLayout(chart_layout)
                
                # 初始化时间记录
                self.start_time = time.time()
                
                # 参数显示区域：名称+固定文本框，平均划分宽度
                field_layout = QHBoxLayout()
                for field in fields:
                    field_widget = QWidget()
                    field_v_layout = QVBoxLayout(field_widget)
                    
                    # 名称标签
                    name_label = QLabel(f"{FIELD_NAMES[field]}")
                    name_label.setAlignment(Qt.AlignCenter)
                    name_label.setStyleSheet("font-size: 14px; font-weight: bold;")
                    
                    # 值显示文本框
                    value_label = QLabel("-")
                    value_label.setAlignment(Qt.AlignCenter)
                    value_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #0066cc; border: 1px solid #ccc; padding: 8px; border-radius: 4px;")
                    value_label.setMinimumHeight(50)
                    
                    field_v_layout.addWidget(name_label)
                    field_v_layout.addWidget(value_label)
                    
                    field_layout.addWidget(field_widget)
                    
                    # 保存引用
                    setattr(self, f"{field}_label", value_label)
                
                category_layout.addLayout(field_layout)
            elif category == "角速度":
                # 角速度标签页：添加图表展示
                chart_layout = QHBoxLayout()
                
                # 创建角速度图表
                self.gyro_chart = QChart()
                self.gyro_chart.setTitle("角速度随时间变化曲线")
                self.gyro_chart.setAnimationOptions(QChart.NoAnimation)  # 禁用动画，提高性能
                
                # 创建三个坐标轴（X轴为时间，Y轴为角速度值）
                self.gyro_axis_x = QValueAxis()
                self.gyro_axis_x.setTitleText("时间 (s)")
                self.gyro_axis_x.setLabelFormat("%.1f")
                self.gyro_axis_x.setRange(0, 10)  # 显示最近10秒数据
                
                self.gyro_axis_y = QValueAxis()
                self.gyro_axis_y.setTitleText("角速度 (°/s)")
                self.gyro_axis_y.setLabelFormat("%.0f")
                self.gyro_axis_y.setRange(-360, 360)  # 合理的角速度范围
                
                # 创建三个系列（X、Y、Z轴）
                self.gyro_series_x = QLineSeries()
                self.gyro_series_x.setName("角速度X (°/s)")
                self.gyro_series_x.setColor(QColor("red"))
                
                self.gyro_series_y = QLineSeries()
                self.gyro_series_y.setName("角速度Y (°/s)")
                self.gyro_series_y.setColor(QColor("green"))
                
                self.gyro_series_z = QLineSeries()
                self.gyro_series_z.setName("角速度Z (°/s)")
                self.gyro_series_z.setColor(QColor("blue"))
                
                # 添加系列和坐标轴到图表
                self.gyro_chart.addSeries(self.gyro_series_x)
                self.gyro_chart.addSeries(self.gyro_series_y)
                self.gyro_chart.addSeries(self.gyro_series_z)
                
                self.gyro_chart.addAxis(self.gyro_axis_x, Qt.AlignBottom)
                self.gyro_chart.addAxis(self.gyro_axis_y, Qt.AlignLeft)
                
                self.gyro_series_x.attachAxis(self.gyro_axis_x)
                self.gyro_series_x.attachAxis(self.gyro_axis_y)
                
                self.gyro_series_y.attachAxis(self.gyro_axis_x)
                self.gyro_series_y.attachAxis(self.gyro_axis_y)
                
                self.gyro_series_z.attachAxis(self.gyro_axis_x)
                self.gyro_series_z.attachAxis(self.gyro_axis_y)
                
                # 创建图表视图
                chart_view = QChartView(self.gyro_chart)
                chart_view.setRenderHint(QPainter.Antialiasing)
                chart_view.setMinimumHeight(300)
                
                chart_layout.addWidget(chart_view)
                category_layout.addLayout(chart_layout)
                
                # 参数显示区域：名称+固定文本框，平均划分宽度
                field_layout = QHBoxLayout()
                for field in fields:
                    field_widget = QWidget()
                    field_v_layout = QVBoxLayout(field_widget)
                    
                    # 名称标签
                    name_label = QLabel(f"{FIELD_NAMES[field]}")
                    name_label.setAlignment(Qt.AlignCenter)
                    name_label.setStyleSheet("font-size: 14px; font-weight: bold;")
                    
                    # 值显示文本框
                    value_label = QLabel("-")
                    value_label.setAlignment(Qt.AlignCenter)
                    value_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #0066cc; border: 1px solid #ccc; padding: 8px; border-radius: 4px;")
                    value_label.setMinimumHeight(50)
                    
                    field_v_layout.addWidget(name_label)
                    field_v_layout.addWidget(value_label)
                    
                    field_layout.addWidget(field_widget)
                    
                    # 保存引用
                    setattr(self, f"{field}_label", value_label)
                
                category_layout.addLayout(field_layout)
            elif category == "姿态":
                # 姿态角标签页：添加图表展示（横向排列的两个图表）
                chart_layout = QHBoxLayout()
                
                # 左侧图表：俯仰角变化曲线
                self.pitch_chart = QChart()
                self.pitch_chart.setTitle("俯仰角随时间变化曲线")
                self.pitch_chart.setAnimationOptions(QChart.NoAnimation)  # 禁用动画，提高性能
                
                # 俯仰角图表坐标轴
                self.pitch_axis_x = QValueAxis()
                self.pitch_axis_x.setTitleText("时间 (s)")
                self.pitch_axis_x.setLabelFormat("%.1f")
                self.pitch_axis_x.setRange(0, 10)  # 显示最近10秒数据
                
                self.pitch_axis_y = QValueAxis()
                self.pitch_axis_y.setTitleText("俯仰角 (°)")
                self.pitch_axis_y.setLabelFormat("%.0f")
                self.pitch_axis_y.setRange(-180, 180)  # 合理的姿态角范围
                
                # 俯仰角系列
                self.pitch_series = QLineSeries()
                self.pitch_series.setName("俯仰角 (°)")
                self.pitch_series.setColor(QColor("red"))
                
                # 添加系列和坐标轴到俯仰角图表
                self.pitch_chart.addSeries(self.pitch_series)
                self.pitch_chart.addAxis(self.pitch_axis_x, Qt.AlignBottom)
                self.pitch_chart.addAxis(self.pitch_axis_y, Qt.AlignLeft)
                
                self.pitch_series.attachAxis(self.pitch_axis_x)
                self.pitch_series.attachAxis(self.pitch_axis_y)
                
                # 创建俯仰角图表视图
                pitch_chart_view = QChartView(self.pitch_chart)
                pitch_chart_view.setRenderHint(QPainter.Antialiasing)
                pitch_chart_view.setMinimumHeight(300)
                
                chart_layout.addWidget(pitch_chart_view)
                
                # 右侧图表：翻滚角变化曲线
                self.roll_chart = QChart()
                self.roll_chart.setTitle("翻滚角随时间变化曲线")
                self.roll_chart.setAnimationOptions(QChart.NoAnimation)  # 禁用动画，提高性能
                
                # 翻滚角图表坐标轴
                self.roll_axis_x = QValueAxis()
                self.roll_axis_x.setTitleText("时间 (s)")
                self.roll_axis_x.setLabelFormat("%.1f")
                self.roll_axis_x.setRange(0, 10)  # 显示最近10秒数据
                
                self.roll_axis_y = QValueAxis()
                self.roll_axis_y.setTitleText("翻滚角 (°)")
                self.roll_axis_y.setLabelFormat("%.0f")
                self.roll_axis_y.setRange(-180, 180)  # 合理的姿态角范围
                
                # 翻滚角系列
                self.roll_series = QLineSeries()
                self.roll_series.setName("翻滚角 (°)")
                self.roll_series.setColor(QColor("green"))
                
                # 添加系列和坐标轴到翻滚角图表
                self.roll_chart.addSeries(self.roll_series)
                self.roll_chart.addAxis(self.roll_axis_x, Qt.AlignBottom)
                self.roll_chart.addAxis(self.roll_axis_y, Qt.AlignLeft)
                
                self.roll_series.attachAxis(self.roll_axis_x)
                self.roll_series.attachAxis(self.roll_axis_y)
                
                # 创建翻滚角图表视图
                roll_chart_view = QChartView(self.roll_chart)
                roll_chart_view.setRenderHint(QPainter.Antialiasing)
                roll_chart_view.setMinimumHeight(300)
                
                chart_layout.addWidget(roll_chart_view)
                
                category_layout.addLayout(chart_layout)
                
                # 参数显示区域：名称+固定文本框，平均划分宽度
                field_layout = QHBoxLayout()
                for field in fields:
                    field_widget = QWidget()
                    field_v_layout = QVBoxLayout(field_widget)
                    
                    # 名称标签
                    name_label = QLabel(f"{FIELD_NAMES[field]}")
                    name_label.setAlignment(Qt.AlignCenter)
                    name_label.setStyleSheet("font-size: 14px; font-weight: bold;")
                    
                    # 值显示文本框
                    value_label = QLabel("-")
                    value_label.setAlignment(Qt.AlignCenter)
                    value_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #0066cc; border: 1px solid #ccc; padding: 8px; border-radius: 4px;")
                    value_label.setMinimumHeight(50)
                    
                    field_v_layout.addWidget(name_label)
                    field_v_layout.addWidget(value_label)
                    
                    field_layout.addWidget(field_widget)
                    
                    # 保存引用
                    setattr(self, f"{field}_label", value_label)
                
                category_layout.addLayout(field_layout)
            elif category == "环境":
                # 环境标签页：添加图表展示（横向排列的两个图表）
                chart_layout = QHBoxLayout()
                
                # 左侧图表：气压变化曲线
                self.pressure_chart = QChart()
                self.pressure_chart.setTitle("气压随时间变化曲线")
                self.pressure_chart.setAnimationOptions(QChart.NoAnimation)  # 禁用动画，提高性能
                
                # 气压图表坐标轴
                self.pressure_axis_x = QValueAxis()
                self.pressure_axis_x.setTitleText("时间 (s)")
                self.pressure_axis_x.setLabelFormat("%.1f")
                self.pressure_axis_x.setRange(0, 10)  # 显示最近10秒数据
                
                self.pressure_axis_y = QValueAxis()
                self.pressure_axis_y.setTitleText("气压 (kPa)")
                self.pressure_axis_y.setLabelFormat("%.1f")
                self.pressure_axis_y.setRange(90, 110)  # 合理的气压范围（90-110 kPa）
                
                # 气压系列
                self.pressure_series = QLineSeries()
                self.pressure_series.setName("气压 (kPa)")
                self.pressure_series.setColor(QColor("red"))
                
                # 添加系列和坐标轴到气压图表
                self.pressure_chart.addSeries(self.pressure_series)
                self.pressure_chart.addAxis(self.pressure_axis_x, Qt.AlignBottom)
                self.pressure_chart.addAxis(self.pressure_axis_y, Qt.AlignLeft)
                
                self.pressure_series.attachAxis(self.pressure_axis_x)
                self.pressure_series.attachAxis(self.pressure_axis_y)
                
                # 创建气压图表视图
                pressure_chart_view = QChartView(self.pressure_chart)
                pressure_chart_view.setRenderHint(QPainter.Antialiasing)
                pressure_chart_view.setMinimumHeight(300)
                
                chart_layout.addWidget(pressure_chart_view)
                
                # 右侧图表：高度变化曲线
                self.altitude_chart = QChart()
                self.altitude_chart.setTitle("高度随时间变化曲线")
                self.altitude_chart.setAnimationOptions(QChart.NoAnimation)  # 禁用动画，提高性能
                
                # 高度图表坐标轴
                self.altitude_axis_x = QValueAxis()
                self.altitude_axis_x.setTitleText("时间 (s)")
                self.altitude_axis_x.setLabelFormat("%.1f")
                self.altitude_axis_x.setRange(0, 10)  # 显示最近10秒数据
                
                self.altitude_axis_y = QValueAxis()
                self.altitude_axis_y.setTitleText("高度 (m)")
                self.altitude_axis_y.setLabelFormat("%.0f")
                self.altitude_axis_y.setRange(-100, 10000)  # 合理的高度范围（-100m到10000m）
                
                # 高度系列
                self.altitude_series = QLineSeries()
                self.altitude_series.setName("高度 (m)")
                self.altitude_series.setColor(QColor("green"))
                
                # 添加系列和坐标轴到高度图表
                self.altitude_chart.addSeries(self.altitude_series)
                self.altitude_chart.addAxis(self.altitude_axis_x, Qt.AlignBottom)
                self.altitude_chart.addAxis(self.altitude_axis_y, Qt.AlignLeft)
                
                self.altitude_series.attachAxis(self.altitude_axis_x)
                self.altitude_series.attachAxis(self.altitude_axis_y)
                
                # 创建高度图表视图
                altitude_chart_view = QChartView(self.altitude_chart)
                altitude_chart_view.setRenderHint(QPainter.Antialiasing)
                altitude_chart_view.setMinimumHeight(300)
                
                chart_layout.addWidget(altitude_chart_view)
                
                category_layout.addLayout(chart_layout)
                
                # 参数显示区域：名称+固定文本框，平均划分宽度
                field_layout = QHBoxLayout()
                for field in fields:
                    field_widget = QWidget()
                    field_v_layout = QVBoxLayout(field_widget)
                    
                    # 名称标签
                    name_label = QLabel(f"{FIELD_NAMES[field]}")
                    name_label.setAlignment(Qt.AlignCenter)
                    name_label.setStyleSheet("font-size: 14px; font-weight: bold;")
                    
                    # 值显示文本框
                    value_label = QLabel("-")
                    value_label.setAlignment(Qt.AlignCenter)
                    value_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #0066cc; border: 1px solid #ccc; padding: 8px; border-radius: 4px;")
                    value_label.setMinimumHeight(50)
                    
                    field_v_layout.addWidget(name_label)
                    field_v_layout.addWidget(value_label)
                    
                    field_layout.addWidget(field_widget)
                    
                    # 保存引用
                    setattr(self, f"{field}_label", value_label)
                
                category_layout.addLayout(field_layout)
            elif category == "位置":
                # 位置标签页：显示地图和经纬度值
                # 地图和参数显示布局
                map_layout = QVBoxLayout()
                
                # 地图显示区域
                self.map_view = QWebEngineView()
                
                # 自定义页面以捕获JavaScript控制台输出
                class MapWebEnginePage(QWebEnginePage):
                    def javaScriptConsoleMessage(self, level, message, line_number, source_id):
                        level_str = str(level)
                        print(f"[Map JS] {level_str} (line {line_number}): {message}")
                
                self.map_view.setPage(MapWebEnginePage())
                
                # 加载高德地图
                map_html = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>位置地图</title>
    <style>
        body { margin: 0; padding: 0; height: 100vh; background-color: #f0f0f0; }
        #map { 
            width: 100%; 
            height: 100%; 
            position: relative;
        }
        .status {
            position: absolute;
            top: 20px;
            left: 20px;
            background: white;
            padding: 15px;
            border-radius: 8px;
            font-family: Arial, sans-serif;
            font-size: 14px;
            z-index: 1000;
            box-shadow: 0 2px 10px rgba(0,0,0,0.2);
        }
        .status.success { color: green; }
        .status.error { color: red; }
        .status.loading { color: orange; }
        .location-log {
            position: absolute;
            top: 20px;
            right: 20px;
            background: white;
            padding: 15px;
            border-radius: 8px;
            font-family: Arial, sans-serif;
            font-size: 12px;
            z-index: 1000;
            box-shadow: 0 2px 10px rgba(0,0,0,0.2);
            max-width: 300px;
            max-height: 150px;
            overflow-y: auto;
        }
        .location-log h4 {
            margin-top: 0;
            color: #333;
            border-bottom: 1px solid #ccc;
            padding-bottom: 5px;
        }
        .location-log .entry {
            margin: 5px 0;
            color: #666;
            border-left: 3px solid #007bff;
            padding-left: 8px;
        }
        .location-log .entry .time {
            font-size: 10px;
            color: #999;
        }
        .current-location {
            position: absolute;
            background: white;
            padding: 8px 12px;
            border-radius: 6px;
            font-family: Arial, sans-serif;
            font-size: 13px;
            font-weight: bold;
            color: #007bff;
            z-index: 1000;
            box-shadow: 0 2px 8px rgba(0,0,0,0.3);
            white-space: nowrap;
            border: 2px solid #007bff;
        }
    </style>
    
    <script type="text/javascript">
        window._AMapSecurityConfig = {
            securityJsCode: "dd1127e11e1f2d5504a2f2ec9824eb78"
        };
    </script>
    <script type="text/javascript" src="https://webapi.amap.com/maps?v=2.0&key=431d3bb1fa78eef96736dc499113fca2"></script>
</head>
<body>
    <div id="map">
        <div class="status loading" id="status">
            <strong>地图加载中...</strong><br>
            <small>正在加载地图资源...</small>
        </div>
        <div class="location-log" id="locationLog">
            <h4>最近位置更新</h4>
        </div>
        <div class="current-location" id="currentLocation">
            位置: 39.90923, 116.397428
        </div>
    </div>

    <script type="text/javascript">
        console.log('Map script loaded');
        
        let map;
        let marker;
        let label; // 坐标标签
        let locationLog = [];
        let lastLocation = null; // 保存上一次坐标
        let apiCallCount = 0; // API调用计数器
        let apiCallStartDate = new Date().toDateString(); // API调用开始日期
        
        // 地球半径，用于计算距离
        const EARTH_RADIUS = 6371000;
        
        function log(msg, isError) {
            const statusDiv = document.getElementById('status');
            const color = isError ? 'red' : 'green';
            statusDiv.innerHTML += `[${new Date().toLocaleTimeString()}] <span style="color:${color}">${msg}</span>\\n`;
            console.log(msg);
        }
        
        function addLocationLog(latitude, longitude) {
            const logEntry = {
                time: new Date().toLocaleTimeString(),
                latitude: latitude,
                longitude: longitude
            };
            
            locationLog.unshift(logEntry);
            
            // 只保留最近3条记录
            if (locationLog.length > 3) {
                locationLog.pop();
            }
            
            // 更新日志显示
            const logDiv = document.getElementById('locationLog');
            const entriesDiv = logDiv.querySelector('.entries') || document.createElement('div');
            entriesDiv.className = 'entries';
            entriesDiv.innerHTML = '';
            
            locationLog.forEach(entry => {
                const entryDiv = document.createElement('div');
                entryDiv.className = 'entry';
                entryDiv.innerHTML = `
                    <span class="time">[${entry.time}]</span><br>
                    ${entry.latitude.toFixed(6)}, ${entry.longitude.toFixed(6)}
                `;
                entriesDiv.appendChild(entryDiv);
            });
            
            // 确保entriesDiv在logDiv中
            if (!logDiv.contains(entriesDiv)) {
                logDiv.appendChild(entriesDiv);
            }
        }
        
        // 计算两个坐标之间的距离（米）
        function calculateDistance(lat1, lon1, lat2, lon2) {
            const dLat = (lat2 - lat1) * Math.PI / 180;
            const dLon = (lon2 - lon1) * Math.PI / 180;
            
            const a = Math.sin(dLat / 2) * Math.sin(dLat / 2) +
                     Math.cos(lat1 * Math.PI / 180) * Math.cos(lat2 * Math.PI / 180) *
                     Math.sin(dLon / 2) * Math.sin(dLon / 2);
            
            const c = 2 * Math.atan2(Math.sqrt(a), Math.sqrt(1 - a));
            
            return EARTH_RADIUS * c;
        }
        
        // 检查API调用限制
        function checkApiCallLimit() {
            const today = new Date().toDateString();
            
            // 如果日期不同，重置计数器
            if (today !== apiCallStartDate) {
                apiCallCount = 0;
                apiCallStartDate = today;
                console.log('API调用计数器已重置（新日期）');
            }
            
            // 检查是否超过限制
            return apiCallCount >= 1000;
        }
        
        // 增加API调用计数
        function incrementApiCallCount() {
            apiCallCount++;
            console.log(`API调用计数: ${apiCallCount}/1000 (${apiCallStartDate})`);
        }
        
        // 显示API调用限制信息
        function showApiLimitInfo() {
            const locationLogDiv = document.getElementById('locationLog');
            const limitDiv = locationLogDiv.querySelector('.api-limit-info') || document.createElement('div');
            limitDiv.className = 'api-limit-info';
            limitDiv.innerHTML = `
                <div style="background: #ffcc00; color: #000; padding: 5px; margin: 5px 0; border-radius: 4px; font-weight: bold; font-size: 11px;">
                    ⚠️ API调用次数超出限制 (${apiCallCount}/1000次)
                </div>
            `;
            
            // 将限制信息添加到日志顶部
            if (!locationLogDiv.contains(limitDiv)) {
                const entriesDiv = locationLogDiv.querySelector('.entries');
                if (entriesDiv) {
                    locationLogDiv.insertBefore(limitDiv, entriesDiv);
                } else {
                    locationLogDiv.appendChild(limitDiv);
                }
            }
        }
        
        function initMap() {
            console.log('Initializing map');
            
            try {
                map = new AMap.Map('map', {
                    zoom: 13,
                    center: [116.397428, 39.90923],
                    viewMode: '2D',
                    mapStyle: 'amap://styles/normal'
                });
                
                log('✅ Map created successfully');
                
                // 添加地图控件（使用正确的方式）
                try {
                    if (typeof AMap.Scale === 'function') {
                        map.addControl(new AMap.Scale({}));
                    }
                    if (typeof AMap.ToolBar === 'function') {
                        map.addControl(new AMap.ToolBar({}));
                    }
                    if (typeof AMap.MapType === 'function') {
                        map.addControl(new AMap.MapType({}));
                    }
                } catch (e) {
                    console.log('地图控件加载失败:', e);
                }
                
                // 创建圆形标记
                marker = new AMap.CircleMarker({
                    center: [116.397428, 39.90923],
                    radius: 8,
                    strokeColor: '#007bff',
                    strokeWeight: 2,
                    fillColor: '#ffffff',
                    fillOpacity: 0.8,
                    map: map
                });
                
                // 创建坐标信息标签
                label = new AMap.Text({
                    text: '39.90923, 116.397428',
                    anchor: 'bottom-left',
                    style: {
                        'background-color': 'white',
                        'border': '1px solid #007bff',
                        'padding': '4px 8px',
                        'border-radius': '4px',
                        'font-size': '12px',
                        'color': '#007bff',
                        'font-weight': 'bold',
                        'white-space': 'nowrap'
                    },
                    position: [116.397428, 39.90923],
                    map: map
                });
                
                // 保存初始位置
                lastLocation = { latitude: 39.90923, longitude: 116.397428 };
                
                document.getElementById('status').className = 'status success';
                document.getElementById('status').innerHTML = `
                    <strong>地图加载成功!</strong><br>
                    <small>高德地图API已准备好</small>
                `;
                
            } catch (error) {
                console.error('Map initialization error:', error);
                document.getElementById('status').className = 'status error';
                document.getElementById('status').innerHTML = `
                    <strong>地图加载失败!</strong><br>
                    <small>错误: ${error.message}</small>
                `;
            }
        }
        
        window.addEventListener('DOMContentLoaded', () => {
            console.log('DOM loaded');
            
            if (typeof AMap !== 'undefined') {
                initMap();
            } else {
                log('❌ AMap API not loaded', true);
            }
        });
        
        // 更新地图位置的函数
        window.updateLocation = function(latitude, longitude) {
            console.log(`Updating location: lat=${latitude}, lon=${longitude}`);
            
            if (map && latitude && longitude) {
                try {
                    // 检查API调用限制
                    if (checkApiCallLimit()) {
                        console.log('API调用次数已超出每日限制');
                        showApiLimitInfo();
                        return;
                    }
                    
                    // 检查距离是否小于指定阈值
                    if (lastLocation) {
                        const distance = calculateDistance(lastLocation.latitude, lastLocation.longitude, latitude, longitude);
                        if (distance < 50) {
                            console.log(`距离较近 (${distance.toFixed(1)}米 < 50米)，不刷新位置`);
                            return;
                        }
                    }
                    
                    // 移除旧标记
                    if (marker) {
                        map.remove(marker);
                    }
                    
                    // 移动地图到新位置
                    map.setZoomAndCenter(15, [longitude, latitude]);
                    
                    // 创建新的圆形标记
                    marker = new AMap.CircleMarker({
                        center: [longitude, latitude],
                        radius: 8,
                        strokeColor: '#007bff',
                        strokeWeight: 2,
                        fillColor: '#ffffff',
                        fillOpacity: 0.8,
                        map: map
                    });
                    
                    // 更新坐标标签
                    if (label) {
                        map.remove(label);
                    }
                    label = new AMap.Text({
                        text: `${latitude.toFixed(6)}, ${longitude.toFixed(6)}`,
                        anchor: 'bottom-left',
                        style: {
                            'background-color': 'white',
                            'border': '1px solid #007bff',
                            'padding': '4px 8px',
                            'border-radius': '4px',
                            'font-size': '12px',
                            'color': '#007bff',
                            'font-weight': 'bold',
                            'white-space': 'nowrap'
                        },
                        position: [longitude, latitude],
                        map: map
                    });
                    
                    // 更新当前位置显示，使其跟随标记
                    const pixelPosition = map.lngLatToContainer([longitude, latitude]);
                    const locationDiv = document.getElementById('currentLocation');
                    locationDiv.textContent = `位置: ${latitude.toFixed(6)}, ${longitude.toFixed(6)}`;
                    locationDiv.style.left = pixelPosition.x + 'px';
                    locationDiv.style.top = (pixelPosition.y - 40) + 'px';
                    locationDiv.style.transform = 'translate(-50%, 0)';
                    
                    // 添加到位置更新日志
                    addLocationLog(latitude, longitude);
                    
                    // 更新上一次位置
                    lastLocation = { latitude, longitude };
                    
                    // 增加API调用计数
                    incrementApiCallCount();
                    
                } catch (error) {
                    console.error('Update location error:', error);
                    log(`❌ Error updating location: ${error.message}`, true);
                }
            }
        };
    </script>
</body>
</html>
                """.strip()
                
                # 替换HTML中的变量值
                map_html = map_html.replace("${MAP_API_DAILY_LIMIT}", str(MAP_API_DAILY_LIMIT))
                map_html = map_html.replace("${MAP_UPDATE_DISTANCE_THRESHOLD}", str(MAP_UPDATE_DISTANCE_THRESHOLD))
                map_html = map_html.replace("${MAP_ZOOM_LEVEL}", str(MAP_ZOOM_LEVEL))
                
                self.map_view.setHtml(map_html)
                self.map_view.setMinimumHeight(400)
                map_layout.addWidget(self.map_view)
                
                category_layout.addLayout(map_layout)
            else:
                # 其他分类继续使用原来的布局
                for field in fields:
                    field_group = QGroupBox(f"{FIELD_NAMES[field]} ({FIELD_UNITS.get(field, '')})")
                    field_layout = QVBoxLayout(field_group)
                    
                    # 实时值显示
                    value_label = QLabel("-")
                    value_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #0066cc;")
                    value_label.setAlignment(Qt.AlignCenter)
                    
                    field_layout.addWidget(value_label)
                    
                    # 保存引用
                    setattr(self, f"{field}_label", value_label)
                    
                    category_layout.addWidget(field_group)
            
            self.tab_widget.addTab(category_tab, category)
        
        parsed_layout.addWidget(self.tab_widget)
        
        content_splitter.addWidget(parsed_widget)
        
        # 设置分割比例
        content_splitter.setSizes([300, 600])
        main_layout.addWidget(content_splitter)
        
        # 底部导出区域 - 紧凑布局
        export_group = QGroupBox("数据导出")
        export_inner_layout = QHBoxLayout(export_group)
        export_inner_layout.setSpacing(8)  # 减少控件间距
        export_inner_layout.setContentsMargins(10, 10, 10, 10)  # 减少边距
        
        # 时间范围选择
        self.start_time_edit = QDateTimeEdit()
        self.start_time_edit.setDateTime(QDateTime.currentDateTime().addDays(-1))
        self.start_time_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.start_time_edit.setFixedHeight(30)  # 减少控件高度
        
        self.end_time_edit = QDateTimeEdit()
        self.end_time_edit.setDateTime(QDateTime.currentDateTime())
        self.end_time_edit.setDisplayFormat("yyyy-MM-dd HH:mm:ss")
        self.end_time_edit.setFixedHeight(30)  # 减少控件高度
        
        # 导出按钮
        self.export_btn = QPushButton("导出Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setFixedHeight(30)  # 减少按钮高度
        self.export_btn.setFixedWidth(100)  # 设置按钮固定宽度
        
        export_inner_layout.addWidget(QLabel("开始时间:"))
        export_inner_layout.addWidget(self.start_time_edit)
        export_inner_layout.addWidget(QLabel("结束时间:"))
        export_inner_layout.addWidget(self.end_time_edit)
        export_inner_layout.addStretch()
        export_inner_layout.addWidget(self.export_btn)
        
        # 直接添加到主布局，不使用分割器，更紧凑
        main_layout.addWidget(export_group)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
    
    def load_subscribed_devices(self):
        """加载已订阅的设备列表"""
        devices = self.db_manager.get_subscribed_devices()
        self.subscribed_devices = devices
        self.device_combo.clear()
        self.device_combo.addItems(devices)
    
    def show_log_window(self):
        """显示日志窗口"""
        self.log_window.show()
        self.log_window.raise_()
        self.log_window.activateWindow()
    
    def redirect_stdout_stderr(self):
        """重定向标准输出和错误到日志窗口"""
        self.original_stdout = sys.stdout
        self.original_stderr = sys.stderr
        
        self.stdout_redirector = StreamRedirector(self.original_stdout)
        self.stderr_redirector = StreamRedirector(self.original_stderr)
        
        sys.stdout = self.stdout_redirector
        sys.stderr = self.stderr_redirector
        
        self.stdout_redirector.output_received.connect(self.log_window.add_log)
        self.stderr_redirector.output_received.connect(self.log_window.add_log)
    
    def subscribe_device(self):
        """订阅/停止订阅设备"""
        if self.subscribe_btn.text() == "订阅设备":
            imei = self.imei_edit.text().strip()
            
            if not imei:
                QMessageBox.warning(self, "警告", "请输入设备IMEI号")
                return
            
            print(f"正在订阅设备: {imei}")
            
            # 停止当前的MQTT连接（不阻塞主线程）
            if self.mqtt_thread:
                self.mqtt_thread.stop()
                # 使用QTimer延迟启动新连接，避免线程冲突
                QTimer.singleShot(1000, lambda: self.start_new_mqtt_thread(imei))
            else:
                self.start_new_mqtt_thread(imei)
            
            # 改变按钮状态
            self.subscribe_btn.setText("停止订阅")
        else:
            # 停止订阅
            print("正在停止订阅")
            if self.mqtt_thread:
                print(f"线程状态: {self.mqtt_thread.isRunning()}")
                
                # 先取消所有信号连接
                self.mqtt_thread.message_received.disconnect()
                self.mqtt_thread.sensor_data_received.disconnect()
                self.mqtt_thread.connection_status.disconnect()
                self.mqtt_thread.error_occurred.disconnect()
                
                # 停止线程
                self.mqtt_thread.stop()
                print("等待线程停止...")
                self.mqtt_thread.wait()
                print(f"线程状态: {self.mqtt_thread.isFinished()}")
                self.mqtt_thread = None
            
            self.subscribe_btn.setText("订阅设备")
            self.status_label.setText("状态: 未连接")
            self.status_label.setStyleSheet("color: gray; font-weight: bold;")
            print("停止订阅完成")
    
    def start_new_mqtt_thread(self, imei):
        """启动新的MQTT线程"""
        try:
            self.mqtt_thread = MqttThread(imei)
            self.mqtt_thread.message_received.connect(self.on_message_received)
            self.mqtt_thread.sensor_data_received.connect(self.on_sensor_data_received)
            self.mqtt_thread.connection_status.connect(self.on_connection_status)
            self.mqtt_thread.error_occurred.connect(self.on_error_occurred)
            self.mqtt_thread.start()
            
            print(f"MQTT线程已启动，正在连接到设备: {imei}")
            
            # 更新设备列表
            if imei not in self.subscribed_devices:
                self.subscribed_devices.append(imei)
                self.device_combo.addItem(imei)
                self.device_combo.setCurrentText(imei)
        except Exception as e:
            print(f"启动MQTT线程失败: {str(e)}")
            QMessageBox.critical(self, "错误", f"启动MQTT线程失败: {str(e)}")
    
    def switch_device(self, imei):
        """切换设备"""
        if imei:
            self.imei_edit.setText(imei)
            # 加载该设备的数据
            self.load_data_from_db()
            # 重新订阅该设备
            self.subscribe_device()
    
    def _init_log_file(self):
        """初始化log文件"""
        # 创建log文件夹（如果不存在）
        log_dir = "log"
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        
        # 获取当前IMEI
        current_imei = self.imei_edit.text().strip()
        
        # 创建log文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if self.log_file_counter > 0:
            self.log_file_path = os.path.join(log_dir, f"{current_imei}_{timestamp}_{self.log_file_counter}.log")
        else:
            self.log_file_path = os.path.join(log_dir, f"{current_imei}_{timestamp}.log")
        
        # 打开log文件
        self.log_file = open(self.log_file_path, 'a', encoding='utf-8')
        self.log_file_size = 0
        print(f"已创建log文件: {self.log_file_path}")
    
    def _check_log_file_size(self):
        """检查log文件大小，超过5MB则创建新文件"""
        if self.log_file_size > self.max_log_file_size:
            self.log_file.close()
            self.log_file_counter += 1
            self._init_log_file()
    
    def on_message_received(self, topic, payload):
        """处理原始消息"""
        # 显示原始数据到系统日志
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
        print(f"[{timestamp}] 主题: {topic}\n内容: {payload}")
        
        # 保存到log文件
        if not self.log_file:
            self._init_log_file()
        
        log_entry = f"[{timestamp}] 主题: {topic}\n内容: {payload}\n\n"
        self.log_file.write(log_entry)
        self.log_file.flush()
        self.log_file_size += len(log_entry.encode('utf-8'))
        
        # 检查log文件大小
        self._check_log_file_size()
    
    def filter_data_by_event(self, event_type):
        """根据事件类型筛选数据"""
        # 显示所有数据或筛选特定事件类型的数据
        for row in range(self.overview_table.rowCount()):
            if event_type == "全部事件":
                self.overview_table.setRowHidden(row, False)
            else:
                event_col = FIELD_ORDER.index('event')
                item = self.overview_table.item(row, event_col)
                if item and item.text() != event_type:
                    self.overview_table.setRowHidden(row, True)
                else:
                    self.overview_table.setRowHidden(row, False)
    
    def on_sensor_data_received(self, data_list):
        """处理解析后的传感器数据"""
        # 保存数据到数据库
        current_imei = self.imei_edit.text().strip()
        self.db_manager.save_data(data_list, current_imei)
        
        # 更新数据总览表
        for data in data_list:
            row = self.overview_table.rowCount()
            self.overview_table.insertRow(row)
            
            for col, field in enumerate(FIELD_ORDER):
                # 处理事件类型字段
                if field == 'event':
                    value = EVENT_TYPES.get(data.get(field, ''), data.get(field, ''))
                elif field == 'pressure':
                    # 气压参数转换为kPa，处理空值
                    pressure_value = data.get(field, '')
                    if pressure_value and pressure_value != '':
                        try:
                            value = float(pressure_value) / 1000.0
                            value = f"{value:.2f}"  # 格式化为两位小数
                        except ValueError:
                            value = ''
                    else:
                        value = ''
                else:
                    # 对于上电包和传感器数据超时事件，除基本字段外，其他字段留空
                    event_type = data.get('event', '')
                    if event_type in ['POWER_ON', 'SENSOR_REPORT_TIMEOUT'] and field not in ['timestamp', 'version', 'packet_order', 'event']:
                        value = ''
                    else:
                        value = data.get(field, '')
                
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.overview_table.setItem(row, col, item)
        
        # 更新分类展示的实时值
        if data_list:
            last_data = data_list[-1]
            for category, fields in FIELD_CATEGORIES.items():
                # 对所有分类（包括加速度、角速度、角度、姿态、环境）都使用单独的字段标签
                for field in fields:
                    # 跳过经纬度字段，因为这些信息已在地图上显示
                    if field in ['longitude', 'latitude']:
                        continue
                        
                    if field in last_data:
                        try:
                            label = getattr(self, f"{field}_label")
                            value = float(last_data[field])
                            # 对于气压字段，需要将Pa转换为kPa
                            if field == 'pressure':
                                value /= 1000.0
                            # 格式化显示
                            if field in ['pressure']:
                                label.setText(f"{value:.2f}")
                            else:
                                label.setText(f"{value:.0f}")
                        except (ValueError, KeyError, AttributeError) as e:
                            print(f"更新字段 {field} 标签时出错: {e}")
            
            # 更新地图位置
            if hasattr(self, 'map_view') and 'latitude' in last_data and 'longitude' in last_data:
                try:
                    latitude = float(last_data['latitude'])
                    longitude = float(last_data['longitude'])
                    # 检查是否是有效的经纬度（合理范围）
                    if -90 <= latitude <= 90 and -180 <= longitude <= 180 and latitude != 0 and longitude != 0:
                        js_code = f"window.updateLocation({latitude}, {longitude})"
                        self.map_view.page().runJavaScript(js_code)
                        # 不再打印坐标更新信息，避免屏幕被占满
                except (ValueError, KeyError) as e:
                    print(f"更新地图位置时出错: {e}")
        
        # 设置图表显示的时间范围
        display_duration = CHART_DISPLAY_DURATION
        current_time = time.time()
        time_diff = current_time - self.start_time
        
        # 更新加速度图表数据
        for data in data_list:
            # 检查是否包含加速度数据
            if 'accel_x' in data and 'accel_y' in data and 'accel_z' in data:
                try:
                    self.accel_series_x.append(time_diff, float(data['accel_x']))
                    self.accel_series_y.append(time_diff, float(data['accel_y']))
                    self.accel_series_z.append(time_diff, float(data['accel_z']))
                except (ValueError, KeyError) as e:
                    print(f"添加加速度图表数据时出错: {e}")
        
        # 更新角速度图表数据
        for data in data_list:
            # 检查是否包含角速度数据
            if 'gyro_x' in data and 'gyro_y' in data and 'gyro_z' in data:
                try:
                    self.gyro_series_x.append(time_diff, float(data['gyro_x']))
                    self.gyro_series_y.append(time_diff, float(data['gyro_y']))
                    self.gyro_series_z.append(time_diff, float(data['gyro_z']))
                except (ValueError, KeyError) as e:
                    print(f"添加角速度图表数据时出错: {e}")
        
        # 更新俯仰角图表数据
        for data in data_list:
            # 检查是否包含俯仰角数据
            if 'attitude1' in data:
                try:
                    self.pitch_series.append(time_diff, float(data['attitude1']))
                except (ValueError, KeyError) as e:
                    print(f"添加俯仰角图表数据时出错: {e}")
        
        # 更新翻滚角图表数据
        for data in data_list:
            # 检查是否包含翻滚角数据
            if 'attitude2' in data:
                try:
                    self.roll_series.append(time_diff, float(data['attitude2']))
                except (ValueError, KeyError) as e:
                    print(f"添加翻滚角图表数据时出错: {e}")
        
        # 更新气压图表数据
        for data in data_list:
            # 检查是否包含气压数据
            if 'pressure' in data:
                try:
                    # 气压值从Pa转换为kPa
                    pressure_kpa = float(data['pressure']) / 1000.0
                    self.pressure_series.append(time_diff, pressure_kpa)
                except (ValueError, KeyError) as e:
                    print(f"添加气压图表数据时出错: {e}")
        
        # 更新高度图表数据
        for data in data_list:
            # 检查是否包含高度数据
            if 'altitude' in data:
                try:
                    self.altitude_series.append(time_diff, float(data['altitude']))
                except (ValueError, KeyError) as e:
                    print(f"添加高度图表数据时出错: {e}")
        
        # 自动平移图表
        if time_diff > display_duration:
            new_x_min = time_diff - display_duration
            new_x_max = time_diff
            
            # 平移所有图表的X轴
            self.accel_axis_x.setRange(new_x_min, new_x_max)
            self.gyro_axis_x.setRange(new_x_min, new_x_max)
            self.pitch_axis_x.setRange(new_x_min, new_x_max)
            self.roll_axis_x.setRange(new_x_min, new_x_max)
            self.pressure_axis_x.setRange(new_x_min, new_x_max)
            self.altitude_axis_x.setRange(new_x_min, new_x_max)
            
            # 移除超出范围的数据点
            self.remove_old_data_points(self.accel_series_x, new_x_min)
            self.remove_old_data_points(self.accel_series_y, new_x_min)
            self.remove_old_data_points(self.accel_series_z, new_x_min)
            
            self.remove_old_data_points(self.gyro_series_x, new_x_min)
            self.remove_old_data_points(self.gyro_series_y, new_x_min)
            self.remove_old_data_points(self.gyro_series_z, new_x_min)
            
            self.remove_old_data_points(self.pitch_series, new_x_min)
            self.remove_old_data_points(self.roll_series, new_x_min)
            self.remove_old_data_points(self.pressure_series, new_x_min)
            self.remove_old_data_points(self.altitude_series, new_x_min)
        
        # 动态调整Y轴范围
        self.adjust_accel_y_axis_range()
        self.adjust_gyro_y_axis_range()
        self.adjust_pitch_y_axis_range()
        self.adjust_roll_y_axis_range()
        self.adjust_pressure_y_axis_range()
        self.adjust_altitude_y_axis_range()
        
        # 应用事件类型筛选
        self.filter_data_by_event(self.event_filter_combo.currentText())
        
        # 自动调整列宽以适应内容
        self.overview_table.resizeColumnsToContents()
        
        # 如果还有空余宽度，则平均分配给各列
        self.adjust_columns_to_fill_space()
        
        # 自动滚动到最新数据所在的行
        if data_list:
            last_row = self.overview_table.rowCount() - 1
            self.overview_table.scrollToItem(self.overview_table.item(last_row, 0))
    
    def adjust_accel_y_axis_range(self):
        """根据当前X轴范围内的加速度数据动态调整Y轴范围"""
        current_x_min = self.accel_axis_x.min()
        current_x_max = self.accel_axis_x.max()
        
        # 收集当前X轴范围内所有数据点的Y值
        all_y_values = []
        
        # 检查X轴范围内的点
        for series in [self.accel_series_x, self.accel_series_y, self.accel_series_z]:
            for i in range(series.count()):
                point = series.at(i)
                if current_x_min <= point.x() <= current_x_max:
                    all_y_values.append(point.y())
        
        if all_y_values:
            # 计算当前范围内的最小值和最大值
            y_min = min(all_y_values)
            y_max = max(all_y_values)
            
            # 计算120%的范围
            range_ = y_max - y_min
            margin = range_ * CHART_Y_AXIS_MARGIN
            
            new_y_min = y_min - margin
            new_y_max = y_max + margin
            
            # 确保范围不会太小（至少有一定的可见区域）
            if new_y_max - new_y_min < 100:
                center = (y_min + y_max) / 2
                new_y_min = center - 50
                new_y_max = center + 50
            
            # 更新Y轴范围
            self.accel_axis_y.setRange(new_y_min, new_y_max)
    
    def adjust_gyro_y_axis_range(self):
        """根据当前X轴范围内的角速度数据动态调整Y轴范围"""
        current_x_min = self.gyro_axis_x.min()
        current_x_max = self.gyro_axis_x.max()
        
        # 收集当前X轴范围内所有数据点的Y值
        all_y_values = []
        
        # 检查X轴范围内的点
        for series in [self.gyro_series_x, self.gyro_series_y, self.gyro_series_z]:
            for i in range(series.count()):
                point = series.at(i)
                if current_x_min <= point.x() <= current_x_max:
                    all_y_values.append(point.y())
        
        if all_y_values:
            # 计算当前范围内的最小值和最大值
            y_min = min(all_y_values)
            y_max = max(all_y_values)
            
            # 计算120%的范围
            range_ = y_max - y_min
            margin = range_ * CHART_Y_AXIS_MARGIN
            
            new_y_min = y_min - margin
            new_y_max = y_max + margin
            
            # 确保范围不会太小（至少有一定的可见区域）
            if new_y_max - new_y_min < 100:
                center = (y_min + y_max) / 2
                new_y_min = center - 50
                new_y_max = center + 50
            
            # 更新Y轴范围
            self.gyro_axis_y.setRange(new_y_min, new_y_max)
    
    def adjust_pitch_y_axis_range(self):
        """根据当前X轴范围内的俯仰角数据动态调整Y轴范围"""
        current_x_min = self.pitch_axis_x.min()
        current_x_max = self.pitch_axis_x.max()
        
        # 收集当前X轴范围内所有数据点的Y值
        all_y_values = []
        
        # 检查X轴范围内的点
        for i in range(self.pitch_series.count()):
            point = self.pitch_series.at(i)
            if current_x_min <= point.x() <= current_x_max:
                all_y_values.append(point.y())
        
        if all_y_values:
            # 计算当前范围内的最小值和最大值
            y_min = min(all_y_values)
            y_max = max(all_y_values)
            
            # 计算120%的范围
            range_ = y_max - y_min
            margin = range_ * CHART_Y_AXIS_MARGIN
            
            new_y_min = y_min - margin
            new_y_max = y_max + margin
            
            # 确保范围不会太小（至少有一定的可见区域）
            if new_y_max - new_y_min < 100:
                center = (y_min + y_max) / 2
                new_y_min = center - 50
                new_y_max = center + 50
            
            # 更新Y轴范围
            self.pitch_axis_y.setRange(new_y_min, new_y_max)
    
    def adjust_roll_y_axis_range(self):
        """根据当前X轴范围内的翻滚角数据动态调整Y轴范围"""
        current_x_min = self.roll_axis_x.min()
        current_x_max = self.roll_axis_x.max()
        
        # 收集当前X轴范围内所有数据点的Y值
        all_y_values = []
        
        # 检查X轴范围内的点
        for i in range(self.roll_series.count()):
            point = self.roll_series.at(i)
            if current_x_min <= point.x() <= current_x_max:
                all_y_values.append(point.y())
        
        if all_y_values:
            # 计算当前范围内的最小值和最大值
            y_min = min(all_y_values)
            y_max = max(all_y_values)
            
            # 计算120%的范围
            range_ = y_max - y_min
            margin = range_ * CHART_Y_AXIS_MARGIN
            
            new_y_min = y_min - margin
            new_y_max = y_max + margin
            
            # 确保范围不会太小（至少有一定的可见区域）
            if new_y_max - new_y_min < 100:
                center = (y_min + y_max) / 2
                new_y_min = center - 50
                new_y_max = center + 50
            
            # 更新Y轴范围
            self.roll_axis_y.setRange(new_y_min, new_y_max)
    
    def adjust_pressure_y_axis_range(self):
        """根据当前X轴范围内的气压数据动态调整Y轴范围"""
        current_x_min = self.pressure_axis_x.min()
        current_x_max = self.pressure_axis_x.max()
        
        # 收集当前X轴范围内所有数据点的Y值
        all_y_values = []
        
        # 检查X轴范围内的点
        for i in range(self.pressure_series.count()):
            point = self.pressure_series.at(i)
            if current_x_min <= point.x() <= current_x_max:
                all_y_values.append(point.y())
        
        if all_y_values:
            # 计算当前范围内的最小值和最大值
            y_min = min(all_y_values)
            y_max = max(all_y_values)
            
            # 计算120%的范围
            range_ = y_max - y_min
            margin = range_ * CHART_Y_AXIS_MARGIN
            
            new_y_min = y_min - margin
            new_y_max = y_max + margin
            
            # 确保范围不会太小（至少有一定的可见区域）
            if new_y_max - new_y_min < 10:
                center = (y_min + y_max) / 2
                new_y_min = center - 5
                new_y_max = center + 5
            
            # 更新Y轴范围
            self.pressure_axis_y.setRange(new_y_min, new_y_max)
    
    def adjust_altitude_y_axis_range(self):
        """根据当前X轴范围内的高度数据动态调整Y轴范围"""
        current_x_min = self.altitude_axis_x.min()
        current_x_max = self.altitude_axis_x.max()
        
        # 收集当前X轴范围内所有数据点的Y值
        all_y_values = []
        
        # 检查X轴范围内的点
        for i in range(self.altitude_series.count()):
            point = self.altitude_series.at(i)
            if current_x_min <= point.x() <= current_x_max:
                all_y_values.append(point.y())
        
        if all_y_values:
            # 计算当前范围内的最小值和最大值
            y_min = min(all_y_values)
            y_max = max(all_y_values)
            
            # 计算120%的范围
            range_ = y_max - y_min
            margin = range_ * CHART_Y_AXIS_MARGIN
            
            new_y_min = y_min - margin
            new_y_max = y_max + margin
            
            # 确保范围不会太小（至少有一定的可见区域）
            if new_y_max - new_y_min < 1000:
                center = (y_min + y_max) / 2
                new_y_min = center - 500
                new_y_max = center + 500
            
            # 更新Y轴范围
            self.altitude_axis_y.setRange(new_y_min, new_y_max)
    
    def load_data_from_db(self):
        """从数据库加载数据并显示"""
        # 清空表格
        self.overview_table.setRowCount(0)
        
        # 获取当前IMEI和时间范围
        current_imei = self.imei_edit.text().strip()
        if not current_imei:
            return
        
        # 查询数据库中的最近指定时间范围数据
        all_data = self.db_manager.query_data(current_imei, 
            QDateTime.currentDateTime().addSecs(-DATA_LOAD_TIME_RANGE), 
            QDateTime.currentDateTime())
        
        # 优化表格数据插入：先设置行数，再批量插入
        self.overview_table.setRowCount(len(all_data))
        
        # 数据库字段索引映射（根据实际表结构）
        db_field_index = {
            'timestamp': 1,
            'version': 2,
            'packet_order': 3,
            'event': 21,
            'accel_x': 4,
            'accel_y': 5,
            'accel_z': 6,
            'gyro_x': 7,
            'gyro_y': 8,
            'gyro_z': 9,
            'angle_x': 10,
            'angle_y': 11,
            'angle_z': 12,
            'attitude1': 13,
            'attitude2': 14,
            'pressure': 15,
            'altitude': 16,
            'longitude': 17,
            'latitude': 18
        }
        
        for row, data in enumerate(all_data):
            for col, field in enumerate(FIELD_ORDER):
                # 处理事件类型字段
                if field == 'event':
                    value = EVENT_TYPES.get(data[db_field_index[field]], data[db_field_index[field]])
                elif field == 'pressure':
                    # 气压参数转换为kPa，处理空值
                    pressure_value = data[db_field_index[field]]
                    if pressure_value and pressure_value != '':
                        try:
                            value = float(pressure_value) / 1000.0
                            value = f"{value:.2f}"  # 格式化为两位小数
                        except ValueError:
                            value = ''
                    else:
                        value = ''
                else:
                    # 对于上电包和传感器数据超时事件，除基本字段外，其他字段留空
                    event_type = data[db_field_index['event']]
                    if event_type in ['POWER_ON', 'SENSOR_REPORT_TIMEOUT'] and field not in ['timestamp', 'version', 'packet_order', 'event']:
                        value = ''
                    else:
                        value = data[db_field_index[field]]
                
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignCenter)
                self.overview_table.setItem(row, col, item)
        
        # 应用事件类型筛选
        self.filter_data_by_event(self.event_filter_combo.currentText())
    
    def on_connection_status(self, status):
        """更新连接状态"""
        self.status_label.setText(f"状态: {status}")
        
        if "成功" in status:
            self.status_label.setStyleSheet("color: green; font-weight: bold;")
        elif "失败" in status or "错误" in status:
            self.status_label.setStyleSheet("color: red; font-weight: bold;")
        else:
            self.status_label.setStyleSheet("color: orange; font-weight: bold;")
    
    def on_error_occurred(self, error):
        """处理错误信息"""
        QMessageBox.critical(self, "错误", error)
        self.status_label.setText(f"状态: 错误 - {error}")
        self.status_label.setStyleSheet("color: red; font-weight: bold;")
    
    def export_to_excel(self):
        """导出数据到Excel"""
        current_imei = self.imei_edit.text().strip()
        
        if not current_imei:
            QMessageBox.warning(self, "警告", "请先选择或订阅一个设备")
            return
        
        # 获取时间范围
        start_time = self.start_time_edit.dateTime()
        end_time = self.end_time_edit.dateTime()
        
        if start_time > end_time:
            QMessageBox.warning(self, "警告", "开始时间不能晚于结束时间")
            return
        
        # 查询数据
        try:
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(30)
            self.progress_bar.setFormat("正在查询数据...")
            
            results = self.db_manager.query_data(
                current_imei, start_time, end_time
            )
            
            if not results:
                QMessageBox.information(self, "提示", "该时间段内没有数据")
                self.progress_bar.setVisible(False)
                return
            
            # 选择保存路径
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存Excel文件", "", "Excel文件 (*.xlsx)"
            )
            
            if not file_path:
                self.progress_bar.setVisible(False)
                return
            
            # 创建Excel文件
            self.progress_bar.setValue(60)
            self.progress_bar.setFormat("正在创建Excel文件...")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "传感器数据"
            
            # 写入表头
            headers = [FIELD_NAMES[f] for f in FIELD_ORDER]
            for col, header in enumerate(headers, 1):
                ws[get_column_letter(col) + '1'] = header
            
            # 数据库字段索引映射（根据实际表结构）
            db_field_index = {
                'timestamp': 1,
                'version': 2,
                'packet_order': 3,
                'event': 21,
                'accel_x': 4,
                'accel_y': 5,
                'accel_z': 6,
                'gyro_x': 7,
                'gyro_y': 8,
                'gyro_z': 9,
                'angle_x': 10,
                'angle_y': 11,
                'angle_z': 12,
                'attitude1': 13,
                'attitude2': 14,
                'pressure': 15,
                'altitude': 16,
                'longitude': 17,
                'latitude': 18
            }
            
            # 写入数据
            for row, data in enumerate(results, 2):
                for col, field in enumerate(FIELD_ORDER, 1):
                    value = data[db_field_index[field]]
                    ws[get_column_letter(col) + str(row)] = value
            
            wb.save(file_path)
            
            self.progress_bar.setValue(100)
            self.progress_bar.setFormat("导出成功")
            
            QMessageBox.information(
                self, "成功", f"数据已成功导出到:\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)
    
    def on_table_scroll(self, value):
        """表格滚动事件处理"""
        # 当滚动到顶部时加载更多数据（之前的5分钟数据）
        if value == 0:
            current_imei = self.imei_edit.text().strip()
            if not current_imei:
                return
                
            # 计算要加载的时间范围
            end_time = self.loaded_earliest_time
            start_time = end_time.addSecs(-DATA_LOAD_TIME_RANGE)
            
            # 查询数据
            new_data = self.db_manager.query_data(current_imei, start_time, end_time)
            
            if new_data:
                # 数据库字段索引映射（根据实际表结构）
                db_field_index = {
                    'timestamp': 1,
                    'version': 2,
                    'packet_order': 3,
                    'event': 21,
                    'accel_x': 4,
                    'accel_y': 5,
                    'accel_z': 6,
                    'gyro_x': 7,
                    'gyro_y': 8,
                    'gyro_z': 9,
                    'angle_x': 10,
                    'angle_y': 11,
                    'angle_z': 12,
                    'attitude1': 13,
                    'attitude2': 14,
                    'pressure': 15,
                    'altitude': 16,
                    'longitude': 17,
                    'latitude': 18
                }
                
                # 在表格顶部插入新数据
                for row, data in enumerate(new_data):
                    self.overview_table.insertRow(row)
                    for col, field in enumerate(FIELD_ORDER):
                        if field == 'event':
                            value = EVENT_TYPES.get(data[db_field_index[field]], data[db_field_index[field]])
                        elif field == 'pressure':
                            # 气压参数转换为kPa，处理空值
                            pressure_value = data[db_field_index[field]]
                            if pressure_value and pressure_value != '':
                                try:
                                    value = float(pressure_value) / 1000.0
                                    value = f"{value:.2f}"  # 格式化为两位小数
                                except ValueError:
                                    value = ''
                            else:
                                value = ''
                        else:
                            event_type = data[db_field_index['event']]
                            if event_type in ['POWER_ON', 'SENSOR_REPORT_TIMEOUT'] and field not in ['timestamp', 'version', 'packet_order', 'event']:
                                value = ''
                            else:
                                value = data[db_field_index[field]]
                        
                        item = QTableWidgetItem(str(value))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.overview_table.setItem(row, col, item)
                
                # 更新已加载的最早时间
                self.loaded_earliest_time = start_time
                
                # 应用事件类型筛选
                self.filter_data_by_event(self.event_filter_combo.currentText())
                
                print(f"已加载更多数据，时间范围: {start_time.toString('yyyy-MM-dd HH:mm:ss')} 到 {end_time.toString('yyyy-MM-dd HH:mm:ss')}")
            else:
                print("没有更多历史数据可加载")
    
    def remove_old_data_points(self, series, min_time):
        """移除图表系列中超出时间范围的数据点（优化：限制最大数据点数量）"""
        # 限制每个图表系列的最大数据点数量，防止图表渲染卡顿
        max_points = CHART_MAX_DATA_POINTS
        
        # 先移除超出时间范围的数据点
        points_to_remove = []
        for i in range(series.count()):
            point = series.at(i)
            if point.x() < min_time:
                points_to_remove.append(i)
        
        # 从后往前移除数据点，避免索引偏移
        for i in reversed(points_to_remove):
            series.remove(i)
        
        # 如果数据点数量仍然超过限制，进一步移除前面的数据点
        if series.count() > max_points:
            points_to_remove = list(range(series.count() - max_points))
            for i in reversed(points_to_remove):
                series.remove(i)
    
    def adjust_columns_to_fill_space(self):
        """调整列宽以填充剩余空间"""
        # 获取表格总宽度（减去滚动条和边框）
        total_width = self.overview_table.viewport().width()
        
        # 计算各列当前宽度总和
        current_total_width = 0
        for col in range(self.overview_table.columnCount()):
            current_total_width += self.overview_table.columnWidth(col)
        
        # 计算剩余空间
        remaining_width = total_width - current_total_width
        
        if remaining_width > 0:
            # 平均分配剩余宽度给各列
            add_width = remaining_width // self.overview_table.columnCount()
            extra = remaining_width % self.overview_table.columnCount()
            
            for col in range(self.overview_table.columnCount()):
                new_width = self.overview_table.columnWidth(col) + add_width
                if col < extra:
                    new_width += 1
                self.overview_table.setColumnWidth(col, new_width)
    
    def load_map_html(self):
        """加载地图HTML内容"""
        map_html = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>位置地图</title>
            <style>
                body {
                    margin: 0;
                    padding: 0;
                }
                #map {
                    width: 100%;
                    height: 100%;
                }
            </style>
            <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" integrity="sha256-p4NxAoJBhIIN+hmNHrzRCf9tD/miZyoHS5obTRR9BMY=" crossorigin=""/>
            <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js" integrity="sha256-20nQCchB9co0qIjJZRGuk2/Z9VM+kNiyxNV1lvTlZBo=" crossorigin=""></script>
        </head>
        <body>
            <div id="map"></div>
            <script>
                // 初始化地图
                var map = L.map('map').setView([30.66, 104.07], 13); // 默认位置：成都
                
                // 添加OpenStreetMap瓦片图层
                L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
                    attribution: '© OpenStreetMap contributors',
                    maxZoom: 19
                }).addTo(map);
                
                // 初始化标记
                var marker = L.marker([30.66, 104.07]).addTo(map);
                marker.bindPopup('默认位置').openPopup();
                
                // 更新地图位置的函数
                window.updateLocation = function(latitude, longitude) {
                    // 移除旧标记
                    map.removeLayer(marker);
                    
                    // 添加新标记
                    marker = L.marker([latitude, longitude]).addTo(map);
                    marker.bindPopup('当前位置: ' + latitude.toFixed(6) + ', ' + longitude.toFixed(6)).openPopup();
                    
                    // 移动地图到新位置
                    map.setView([latitude, longitude], 15);
                };
            </script>
        </body>
        </html>
        """
        
        # 加载HTML到QWebEngineView
        self.map_view.setHtml(map_html)
    
    def update_map_location(self, latitude, longitude):
        """更新地图位置"""
        # 检查是否是有效坐标
        if latitude is not None and longitude is not None and latitude != 0 and longitude != 0:
            # 使用JavaScript更新地图
            js_code = f"window.updateLocation({latitude}, {longitude})"
            self.map_view.page().runJavaScript(js_code)
    
    
    def closeEvent(self, event):
        """关闭事件处理"""
        if self.mqtt_thread:
            self.mqtt_thread.stop()
            self.mqtt_thread.wait()
        
        # 关闭log文件
        if self.log_file:
            self.log_file.close()
        
        event.accept()

# =============================================================================
# 主函数
# =============================================================================
def main():
    """主函数"""
    app = QApplication(sys.argv)
    
    # 设置应用程序样式
    app.setStyle("Fusion")
    
    # 设置应用程序图标
    icon_path = "dwdl_logo.png"
    if os.path.exists(icon_path):
        app.setWindowIcon(QIcon(QPixmap(icon_path)))
        print(f"已加载应用程序图标: {icon_path}")
    else:
        print(f"警告: 图标文件 {icon_path} 未找到")
    
    # 创建主窗口
    window = MainWindow()
    window.show()
    
    # 运行应用程序
    sys.exit(app.exec())

if __name__ == "__main__":
    main()