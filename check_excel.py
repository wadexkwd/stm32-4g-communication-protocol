#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""检查Excel文件内容"""

from openpyxl import load_workbook

EXCEL_FILE = "sensor_data.xlsx"

def check_excel():
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        print("工作表标题:", ws.title)
        print("表头:")
        for col in range(1, ws.max_column + 1):
            print(f"  {ws.cell(row=1, column=col).value}")
        print(f"数据行数: {ws.max_row - 1}")
        
        if ws.max_row > 1:
            print("\n前3条数据:")
            for row in range(2, min(ws.max_row + 1, 5)):
                data = []
                for col in range(1, ws.max_column + 1):
                    data.append(str(ws.cell(row=row, column=col).value))
                print(f"  行{row - 1}: {', '.join(data)}")
                
    except Exception as e:
        print(f"❌ 检查Excel文件失败: {e}")

if __name__ == "__main__":
    check_excel()
